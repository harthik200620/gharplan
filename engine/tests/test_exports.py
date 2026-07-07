"""Exporter smoke tests — produce valid DXF / XLSX / PDF bytes."""

from __future__ import annotations

import io

import ezdxf
from openpyxl import load_workbook

from app.exporters.dxf import build_dxf
from app.exporters.pdf import build_pdf
from app.exporters.xlsx import build_xlsx
from app.models.enums import City, FinishTier
from app.models.export import Branding
from app.services.boq_service import generate_boq
from app.services.code_service import check_code
from app.services.plan_service import normalize
from app.services.rates import get_rates_provider
from app.services.rules import get_boq_rules, get_code_rules, get_vastu_rules
from app.services.vastu_service import check_vastu


def _boq(plan):
    return generate_boq(plan, City.Bengaluru, FinishTier.standard, get_rates_provider(), get_boq_rules())


def test_dxf_is_valid_r2010(sample_plan):
    plan, _ = normalize(sample_plan)
    data = build_dxf(plan)
    assert len(data) > 500
    doc = ezdxf.read(io.StringIO(data.decode("utf-8")))
    polylines = [e for e in doc.modelspace() if e.dxftype() == "LWPOLYLINE"]
    # one per room + plot boundary (+ north arrow shapes)
    assert len(polylines) >= len(plan.rooms) + 1


def test_xlsx_has_branding_and_rows(sample_plan):
    plan, _ = normalize(sample_plan)
    data = build_xlsx(_boq(plan), Branding(studio_name="Acme Interiors"))
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert ws["A1"].value == "Acme Interiors"
    assert ws.max_row > 10  # header + many BOQ lines + totals


def test_pdf_starts_with_magic_and_has_size(sample_plan):
    plan, _ = normalize(sample_plan)
    v = check_vastu(plan, get_vastu_rules())
    c = check_code(plan, get_code_rules())
    b = _boq(plan)
    data = build_pdf(plan, v, c, b, Branding(studio_name="Acme Studio", gstin="29ABCDE1234F1Z5"))
    assert data[:4] == b"%PDF"
    assert len(data) > 3000


# --------------------------------------------------------------------------- #
# Structural annexe + municipal title block
# --------------------------------------------------------------------------- #


def test_pdf_with_structural_has_design_basis_and_signoff(sample_plan):
    from app.structural import design_structure

    plan, _ = normalize(sample_plan)
    v = check_vastu(plan, get_vastu_rules())
    c = check_code(plan, get_code_rules())
    s = design_structure(plan)
    data = build_pdf(plan, v, c, _boq(plan), Branding(studio_name="Acme Studio"), structural=s)
    assert data[:4] == b"%PDF"
    assert b"Structural Design Basis" in data
    assert b"Sign-off" in data


def test_dxf_with_structural_has_struct_layers(sample_plan):
    from app.structural import design_structure

    plan, _ = normalize(sample_plan)
    s = design_structure(plan)
    data = build_dxf(plan, structural=s)
    assert b"STRUCT-COL" in data
    assert b"STRUCT-GRID" in data
    assert b"STRUCT-FOOTING" in data
