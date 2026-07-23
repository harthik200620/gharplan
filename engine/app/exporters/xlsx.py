"""XLSX export — BOQ plus working-drawing schedules (openpyxl).

The first/active sheet is always the BOQ. When a plan (+ code report) is given,
extra sheets carry the door/window schedule, finishes, area statement and the
MEP coordination summary — the same data the PDF and on-screen viewer show.
"""

from __future__ import annotations

import io
from collections import Counter
from typing import Optional, TYPE_CHECKING

if TYPE_CHECKING:  # optional structural annexe — no hard runtime dependency
    from app.structural.models import StructuralDesign

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.models.boq import BoqReport
from app.models.export import Branding
from app.models.enums import room_label
from app.models.plan import Plan
from app.models.reports import CodeReport
from app.services import schedules as sched
from app.services.mep_model import build_mep_model

_HEADERS = [
    ("Room", 22),
    ("Trade", 16),
    ("Item Code", 12),
    ("Description", 38),
    ("Unit", 7),
    ("Qty", 10),
    ("Material", 11),
    ("Labour", 11),
    ("Rate", 11),
    ("Amount", 13),
    ("HSN", 9),
    ("GST%", 7),
    ("GST Amt", 12),
    ("Total", 14),
]

_HEAD_FILL = PatternFill("solid", fgColor="1F3A5F")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_BOLD = Font(bold=True)
_MONEY = "#,##0.00"
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header(ws: Worksheet, columns: list[tuple[str, int]], row: int = 1) -> int:
    for col, (name, width) in enumerate(columns, start=1):
        c = ws.cell(row, col, name)
        c.fill = _HEAD_FILL
        c.font = _HEAD_FONT
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    return row + 1


def _schedule_sheets(wb: Workbook, plan: Plan, code: Optional[CodeReport], tier: str = "standard") -> None:
    # --- Door & Window (joinery schedule — GFC-07) ---
    ws = wb.create_sheet("Door & Window")
    row = _header(
        ws,
        [
            ("Mark", 8), ("Type", 10), ("Type detail", 20), ("Description", 20),
            ("Width (mm)", 11), ("Height (mm)", 11), ("Qty", 6),
            ("Frame material", 24), ("Glazing / panel", 20), ("Hardware", 32),
            ("U-value", 12), ("SHGC", 8),
        ],
    )
    for g in sched.opening_schedule(plan, tier):
        for col, v in enumerate(
            [
                g.mark, sched.type_label(g), g.type_detail, g.description,
                sched.to_mm(g.width_m), sched.to_mm(g.height_m), g.qty,
                g.frame_material, g.glazing, g.hardware, g.u_value, g.shgc,
            ],
            start=1,
        ):
            ws.cell(row, col, v).border = _BORDER
        row += 1

    # --- Finishes ---
    ws = wb.create_sheet("Finishes")
    row = _header(ws, [("Space", 22), ("Floor", 20), ("Skirting / Dado", 26), ("Walls", 22), ("Ceiling", 18)])
    for tp in sched.present_types(plan):
        fin = sched.finish_for(tp)
        for col, v in enumerate([room_label(tp), fin.floor, fin.dado, fin.walls, fin.ceiling], start=1):
            ws.cell(row, col, v).border = _BORDER
        row += 1

    # --- RCP ceiling treatment (GFC-08 reference) ---
    row += 2
    ws.cell(row, 1, "Reflected ceiling plan — treatment & drop").font = _BOLD
    row = _header(ws, [("Space", 22), ("Treatment", 20), ("Drop (mm)", 12)], row=row + 1)
    for tp in sched.present_types(plan):
        ct = sched.ceiling_treatment_for(tp, tier)
        for col, v in enumerate([room_label(tp), ct.label, ct.drop_mm or "—"], start=1):
            ws.cell(row, col, v).border = _BORDER
        row += 1

    # --- Area statement ---
    ws = wb.create_sheet("Area Statement")
    row = _header(ws, [("Item", 26), ("Metric", 22), ("Imperial", 22)])
    if code is not None:
        rows = sched.area_statement(plan, code.metrics)
    else:
        rows = [{"label": "Plot area", "metric": f"{plan.plot.area_sqm:.1f} m2", "imperial": f"{sched.sqft(plan.plot.area_sqm)} ft2"}]
    for r in rows:
        for col, v in enumerate([r["label"], r["metric"], r["imperial"]], start=1):
            ws.cell(row, col, v).border = _BORDER
        row += 1
    pf = sched.per_floor_built_up(plan)
    if pf:
        for fl, sqm in pf:
            for col, v in enumerate([f"{sched.floor_name(fl)} built-up", f"{sqm:.1f} m2", f"{sched.sqft(sqm)} ft2"], start=1):
                ws.cell(row, col, v).border = _BORDER
            row += 1

    # --- MEP & clashes ---
    ws = wb.create_sheet("MEP & Clashes")
    m = build_mep_model(plan)
    ws.cell(1, 1, "MEP coordination summary").font = _BOLD
    ws.cell(2, 1, f"Fixtures: {len(m.fixtures)}   Pipe runs: {len(m.pipes)}   Electrical points: {len(m.elec)}")
    ws.cell(3, 1, f"Clashes: {m.summary['errors']} errors, {m.summary['warns']} warnings")
    row = _header(ws, [("Severity", 12), ("Rule", 20), ("Issue", 52)], row=5)
    for cl in m.clashes:
        for col, v in enumerate([cl.severity.upper(), cl.rule_id, cl.message], start=1):
            ws.cell(row, col, v).border = _BORDER
        row += 1
    if not m.clashes:
        ws.cell(row, 1, "No MEP coordination clashes detected.")
        row += 1

    # --- circuit / load schedule ---
    row += 2
    ws.cell(row, 1, "Circuit schedule").font = _BOLD
    row = _header(ws, [("Circuit", 22), ("MCB", 8), ("Phase", 8), ("Wire mm2", 10), ("Points", 8)], row=row + 1)
    for ck in m.circuits:
        for col, v in enumerate([ck.name, f"{ck.mcb_a} A", ck.phase, f"{ck.wire_sqmm:g}", ck.points], start=1):
            ws.cell(row, col, v).border = _BORDER
        row += 1
    s = m.summary
    ws.cell(
        row, 1,
        f"Connected load: {s.get('connectedLoadKw', '?')} kW    "
        f"Demand (x{s.get('diversityFactor', 0.6)}): {s.get('demandLoadKw', '?')} kW    "
        f"Recommended service: {s.get('recommendedService', '?')}",
    )

    # --- fixture schedule ---
    row += 3
    ws.cell(row, 1, "Fixture schedule").font = _BOLD
    row = _header(ws, [("Room", 22), ("Fixture", 20)], row=row + 1)
    grouped = Counter((f.room_id, f.kind) for f in m.fixtures)
    for (rid, kind), cnt in sorted(grouped.items()):
        for col, v in enumerate([rid, f"{kind} x{cnt}" if cnt > 1 else kind], start=1):
            ws.cell(row, col, v).border = _BORDER
        row += 1


def _structural_sheet(wb: Workbook, structural: "StructuralDesign") -> None:
    """Preliminary RCC design annexe — the same data the PDF's structural section and
    the on-screen Structure tab show, but untruncated (no page-length limit here)."""
    ws = wb.create_sheet("Structural")
    seismic_zone = (structural.seismic or {}).get("zone", "—")
    ws.cell(1, 1, "Structural design basis (preliminary)").font = Font(bold=True, size=14)
    ws.cell(
        2, 1,
        f"{structural.concrete_grade} concrete · {structural.steel_grade} steel · "
        f"SBC {structural.sbc_kpa:g} kPa ({structural.soil_type}) · Seismic zone {seismic_zone}",
    )
    row = 4

    ws.cell(row, 1, "Grid lines").font = _BOLD
    row = _header(ws, [("Axis", 8), ("Label", 10), ("Offset (m)", 12)], row=row + 1)
    for g in sorted(structural.grid, key=lambda gl: (gl.axis, gl.offset_m)):
        for col, v in enumerate([g.axis.upper(), g.label, round(g.offset_m, 2)], start=1):
            ws.cell(row, col, v).border = _BORDER
        row += 1

    row += 2
    ws.cell(row, 1, f"Member schedule ({len(structural.members)} members)").font = _BOLD
    row = _header(
        ws,
        [("Member", 10), ("Kind", 12), ("Floor", 7), ("Size (mm)", 14), ("Rebar", 42), ("Utilization", 11), ("Clause refs", 30)],
        row=row + 1,
    )
    kind_order = {"column": 0, "footing": 1, "plinth_beam": 2, "beam": 3, "slab": 4, "lintel": 5}
    for mem in sorted(structural.members, key=lambda mm: (kind_order.get(mm.kind, 9), mm.id)):
        size = f"{mem.size_mm[0]}×{mem.size_mm[1]}" + (f" / {mem.thickness_mm} thk" if mem.thickness_mm else "")
        for col, v in enumerate(
            [mem.id, mem.kind.replace("_", " "), mem.floor, size, mem.rebar, mem.utilization, ", ".join(mem.clause_refs)],
            start=1,
        ):
            c = ws.cell(row, col, v)
            c.border = _BORDER
            if col == 6:
                c.number_format = "0%"
        row += 1

    row += 2
    ws.cell(row, 1, f"Bar-bending schedule ({len(structural.bbs)} rows)").font = _BOLD
    row = _header(
        ws,
        [("Mark", 10), ("Member", 10), ("Dia (mm)", 9), ("Shape", 10), ("Count", 8), ("Cut length (m)", 14), ("Weight (kg)", 12)],
        row=row + 1,
    )
    for b in structural.bbs:
        for col, v in enumerate(
            [b.mark, b.member_id, b.bar_dia_mm, b.shape, b.count, round(b.cut_length_m, 2), round(b.total_kg, 2)], start=1
        ):
            ws.cell(row, col, v).border = _BORDER
        row += 1
    row += 1
    ws.cell(row, 1, f"Total steel: {sum(b.total_kg for b in structural.bbs):.0f} kg").font = _BOLD

    row += 3
    ws.cell(row, 1, "Design basis").font = _BOLD
    row += 1
    for sec in structural.design_basis:
        ws.cell(row, 1, sec.title).font = Font(bold=True, italic=True)
        row += 1
        ws.cell(row, 1, sec.body)
        row += 1
        if sec.clause_refs:
            ws.cell(row, 1, "Refs: " + ", ".join(sec.clause_refs)).font = Font(size=8, color="888888")
            row += 1
        row += 1

    ws.cell(row + 1, 1, structural.disclaimer).font = Font(italic=True, size=8, color="888888")


def build_xlsx(
    report: BoqReport,
    branding: Branding | None = None,
    *,
    plan: Optional[Plan] = None,
    code: Optional[CodeReport] = None,
    structural: "StructuralDesign | None" = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    row = 1
    studio = branding.studio_name if branding else "Vastukala AI"
    ws.cell(row, 1, studio).font = Font(bold=True, size=14)
    row += 1
    ws.cell(row, 1, f"Bill of Quantities — {report.city.value} — {report.finish_tier.value.title()} finish")
    row += 2

    row = _header(ws, _HEADERS, row=row)

    for ln in report.lines:
        values = [
            ln.room_label or "",
            ln.trade,
            ln.item_code,
            ln.description,
            ln.unit,
            ln.qty,
            float(ln.material_rate),
            float(ln.labour_rate),
            float(ln.rate),
            float(ln.amount),
            ln.hsn_code,
            float(ln.gst_percent),
            float(ln.gst_amount),
            float(ln.total),
        ]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row, col, v)
            c.border = _BORDER
            if col in (6, 7, 8, 9, 10, 13, 14):
                c.number_format = _MONEY
        row += 1

    # totals block
    row += 1
    s = report.summary
    for label, value in [
        ("Subtotal", float(s.subtotal)),
        ("CGST", float(s.cgst_total)),
        ("SGST", float(s.sgst_total)),
        ("Total GST", float(s.gst_total)),
        ("Grand Total", float(s.grand_total)),
    ]:
        ws.cell(row, 13, label).font = _BOLD
        c = ws.cell(row, 14, value)
        c.font = _BOLD
        c.number_format = _MONEY
        row += 1

    ws.cell(row + 1, 1, report.disclaimer).font = Font(italic=True, size=8, color="888888")

    if plan is not None:
        _schedule_sheets(wb, plan, code, report.finish_tier.value)
    if structural is not None:
        _structural_sheet(wb, structural)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
