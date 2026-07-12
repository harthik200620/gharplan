"""The architect's sheet set — elevations, section, MEP and schedules are
projected from the Plan and carried into the DXF / XLSX / PDF exports."""

from __future__ import annotations

import io

import ezdxf
from openpyxl import load_workbook

from app.exporters.dxf import build_dxf
from app.exporters.xlsx import build_xlsx
from app.models.enums import City, FinishTier
from app.models.export import Branding
from app.services.boq_service import generate_boq
from app.services.code_service import check_code
from app.services.elevations import elevation_openings, roof_level, section_model
from app.services.mep_model import build_mep_model
from app.services.plan_service import normalize
from app.services.rates import get_rates_provider
from app.services.rules import get_boq_rules, get_code_rules
from app.services.schedules import area_statement, opening_schedule


def _boq(plan):
    return generate_boq(plan, City.Bengaluru, FinishTier.standard, get_rates_provider(), get_boq_rules())


def test_drawing_models_project_from_plan(sample_plan):
    plan, _ = normalize(sample_plan)
    # a positive roof level and a section that cuts at least one room
    assert roof_level(plan) > 0
    sm = section_model(plan)
    assert sm.cut_axis in ("x", "y")
    assert len(sm.cells) >= 1
    # the front elevation carries the entrance door
    front = "E"  # sample is East-facing
    door = [o for o in elevation_openings(plan, front, front) if o.kind == "door"]
    assert len(door) == 1
    # MEP derives fixtures + a single DB and finds the (faithful) clashes
    m = build_mep_model(plan)
    assert m.fixtures and m.db is not None
    assert m.summary["errors"] + m.summary["warns"] == len(m.clashes)
    # schedules
    assert any(g.mark.startswith("D") for g in opening_schedule(plan))


def test_dxf_carries_every_view(sample_plan):
    plan, _ = normalize(sample_plan)
    code = check_code(plan, get_code_rules())
    data = build_dxf(plan, code)
    doc = ezdxf.read(io.StringIO(data.decode("latin-1")))
    layers = {layer.dxf.name for layer in doc.layers}
    for expected in ("ELEV", "SECTION", "SECTION_POCHE", "MEP_SOIL", "MEP_ELEC", "MEP_NODE", "SCHEDULE"):
        assert expected in layers, f"missing DXF layer {expected}"
    kinds = {e.dxftype() for e in doc.modelspace()}
    assert {"LWPOLYLINE", "TEXT", "MTEXT", "CIRCLE"} <= kinds


def test_xlsx_has_schedule_sheets(sample_plan):
    plan, _ = normalize(sample_plan)
    code = check_code(plan, get_code_rules())
    data = build_xlsx(_boq(plan), Branding(studio_name="Acme"), plan=plan, code=code)
    wb = load_workbook(io.BytesIO(data))
    assert wb.active.title == "BOQ"  # BOQ stays the first/active sheet
    for sheet in ("Door & Window", "Finishes", "Area Statement", "MEP & Clashes"):
        assert sheet in wb.sheetnames


def test_xlsx_mep_sheet_carries_circuit_and_fixture_schedules(sample_plan):
    """The MEP sheet is no longer counts-only: it tabulates the DB circuit schedule
    (with connected/demand load) and a room-by-room fixture schedule."""
    plan, _ = normalize(sample_plan)
    code = check_code(plan, get_code_rules())
    wb = load_workbook(io.BytesIO(build_xlsx(_boq(plan), Branding(), plan=plan, code=code)))
    ws = wb["MEP & Clashes"]
    vals = {c.value for row in ws.iter_rows() for c in row if isinstance(c.value, str)}
    assert "Circuit schedule" in vals
    assert "Fixture schedule" in vals
    assert any(v and "Lighting" in v for v in vals)  # at least the lighting circuit row
    assert any(v and "Connected load" in v for v in vals)  # the load summary line


def test_area_statement_matches_code_metrics(sample_plan):
    plan, _ = normalize(sample_plan)
    code = check_code(plan, get_code_rules())
    rows = area_statement(plan, code.metrics)
    labels = [r["label"] for r in rows]
    assert labels == ["Plot area", "Built-up area", "Ground coverage", "FAR (used / allowed)", "Number of floors"]


def test_dxf_wiring_on_separate_conduit_layers(sample_plan):
    """Sub-mains, switch-legs and dedicated radials each land on their own CAD layer,
    and switch-legs (board → every light / fan / socket) outnumber the sub-mains."""
    from collections import Counter

    plan, _ = normalize(sample_plan)
    code = check_code(plan, get_code_rules())
    doc = ezdxf.read(io.StringIO(build_dxf(plan, code).decode("latin-1")))
    counts = Counter(e.dxf.layer for e in doc.modelspace() if e.dxf.layer.startswith("MEP_CONDUIT"))
    assert counts["MEP_CONDUIT_SUBMAIN"] > 0
    assert counts["MEP_CONDUIT_DEDICATED"] > 0
    assert counts["MEP_CONDUIT_SWITCHLEG"] > counts["MEP_CONDUIT_SUBMAIN"]


def test_pdf_ships_the_mep_section(sample_plan, monkeypatch):
    """The client PDF now renders the MEP services sheets (previously the MepFlowable
    render code was orphaned and never added to the story). build_mep_model is
    invoked once per MepFlowable + the legend, so a nonzero call count proves the
    section reached the document."""
    from app.exporters import pdf as pdf_mod
    from app.services.rules import get_vastu_rules
    from app.services.vastu_service import check_vastu

    plan, _ = normalize(sample_plan)
    calls = {"n": 0}
    real = pdf_mod.build_mep_model

    def spy(*a, **k):
        calls["n"] += 1
        return real(*a, **k)

    monkeypatch.setattr(pdf_mod, "build_mep_model", spy)
    out = pdf_mod.build_pdf(
        plan,
        check_vastu(plan, get_vastu_rules()),
        check_code(plan, get_code_rules()),
        _boq(plan),
    )
    assert out[:5] == b"%PDF-"
    assert calls["n"] >= 3  # electrical + plumbing sheet + services legend
